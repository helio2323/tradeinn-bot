import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from PIL import Image as PILImage
from src.models.Sqclass import Sqclass
from tqdm import tqdm

def save_catalog(lst_id):
    bd = Sqclass()

    data = bd.get_prod_xlsx(id_list=lst_id)  # Obtém os dados com a função ajustada

    prod_list = bd.get_product_list()
    
    for prod in prod_list:
        if prod[0] == lst_id:
            name_list = prod[1]
            break

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"

    # Adiciona cabeçalhos
    headers = ["ID", "Descrição", "Foto", "Valor R$"]
    ws1.append(headers)

    # Pasta temporária para armazenar imagens
    temp_image_folder = './temp_images'
    os.makedirs(temp_image_folder, exist_ok=True)

    # Adiciona a barra de progresso
    for row in tqdm(data, desc="Processing data", unit="row"):
        product_id, title, image_binary, price_b2b = row
        
        # Verifique se price_b2b não está vazio ou nulo
       # print(f"Processing: ID={product_id}, Title={title}, Price B2B={price_b2b}")

        # Salva a imagem temporariamente
        image_path = None
        if image_binary:
            image_path = os.path.join(temp_image_folder, f'{product_id}.png')
            with open(image_path, 'wb') as image_file:
                image_file.write(image_binary)

            # Redimensiona a imagem
            with PILImage.open(image_path) as img:
                img = img.resize((150, 150))  # Ajuste o tamanho conforme necessário
                img.save(image_path)

        # Adiciona os dados ao Excel
        excel_row = [product_id, title]  # Adiciona ID e Title
        if image_binary:
            # Adiciona a imagem na coluna correta
            img = ExcelImage(image_path)
            img.width = 150  # Ajuste a largura da imagem conforme necessário
            img.height = 150  # Ajuste a altura da imagem conforme necessário
            ws1.add_image(img, f'C{ws1.max_row + 1}')  # Adiciona a imagem na coluna C
        else:
            excel_row.append("No Image")

        # Adiciona o preço B2B à linha
        excel_row.append(1)
        excel_row.append(price_b2b)
        ws1.append(excel_row)
        
        # Ajusta a altura da linha
        ws1.row_dimensions[ws1.max_row].height = 120  # Ajuste a altura da linha conforme necessário
        
        
        # Ajusta largura de todas colunas
        for col in range(1, ws1.max_column + 1):
            ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = 20

        # Centraliza o conteúdo das células
        for row in range(1, ws1.max_row + 1):
            for col in range(1, ws1.max_column + 1):
                ws1.cell(row=row, column=col).alignment = Alignment(vertical='center', horizontal='center')

    ws1.column_dimensions['A'].width = 15  # Ajusta a largura da coluna 'A'
    ws1.column_dimensions['B'].width = 50  # Ajusta a largura da coluna 'B'
   
    save_directory = './Catalogo'
    # Garante que o diretório de salvamento exista
    os.makedirs(save_directory, exist_ok=True)
    
    # Limpa caracteres inválidos do nome do arquivo
    safe_name_list = "".join(c for c in name_list if c.isalnum() or c in [' ', '-'])
    arquive_name = os.path.join(save_directory, f'{lst_id}-{safe_name_list}.xlsx')

    # Salva o arquivo Excel
    wb.save(filename=arquive_name)

    # Remove imagens temporárias
    for file in os.listdir(temp_image_folder):
        os.remove(os.path.join(temp_image_folder, file))
    os.rmdir(temp_image_folder)



import os
from jinja2 import Template
from src.models.Sqclass import Sqclass
#import weasyprint

def generate_html_from_db(id_list, html_file):
    bd = Sqclass()

    # Obtém os dados do banco de dados
    data = bd.get_prod_pdf(id_list)

    # Define o template HTML
    html_template = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Product Catalog</title>
        <style>
            table {
                width: 100%;
                border-collapse: collapse;
            }
            table, th, td {
                border: 1px solid black;
            }
            th, td {
                padding: 8px;
                text-align: left;
            }
            img {
                width: 150px;
                height: 150px;
            }
        </style>
    </head>
    <body>
        <h1>Product Catalog</h1>
        <table>
            <thead>
                <tr>
                    <th>ID</th>
                    <th>Descrição</th>
                    <th>Foto</th>
                    <th>Valor</th>
                </tr>
            </thead>
            <tbody>
                {% for row in data %}
                <tr>
                    <td>{{ row[0] }}</td>
                    <td>{{ row[1] }}</td>
                    <td><img src="{{ row[2] }}" alt="Product Image"></td>
                    <td>{{ row[3] }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </body>
    </html>
    """

    # Cria o template Jinja2
    template = Template(html_template)
    
    # Renderiza o HTML com os dados do banco
    html_content = template.render(data=data)
    
    # Garante que o diretório de salvamento exista
    save_directory = os.path.dirname(html_file)
    print(save_directory)
    os.makedirs(save_directory, exist_ok=True)
    
    # Salva o HTML em um arquivo
    try:
        with open(html_file, 'w') as file:
            file.write(html_content)
    except IOError as e:
        print(f"Erro ao salvar o arquivo HTML: {e}")
        raise

"""def convert_html_to_pdf(html_file, pdf_file):
    print("Convertendo HTML para PDF... aguarde...")
    try:
        weasyprint.HTML(html_file).write_pdf(pdf_file)
        print('PDF gerado com sucesso!!')
    except Exception as e:
        print(f"Erro ao converter HTML para PDF: {e}")
        raise"""

def generate_pdf_from_db(id_list, html_file, pdf_file):
    try:
        generate_html_from_db(id_list, html_file)
       # convert_html_to_pdf(html_file, pdf_file)
    except Exception as e:
        print(f"Erro ao gerar o PDF: {e}")
        raise

